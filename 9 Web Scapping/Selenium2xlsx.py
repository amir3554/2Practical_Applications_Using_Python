import openpyxl.writer
import selenium
import selenium.webdriver
from selenium.webdriver.common.by import By
import openpyxl
from pathlib import Path
import time
import chromedriver_autoinstaller


# https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers
chromedriver_autoinstaller.install()
browser = selenium.webdriver.Chrome()
browser.get('https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers')
time.sleep(5)


all_tables =browser.find_elements(By.TAG_NAME, 'table')


checked_tables_caption = []
for table in all_tables:
       if table.find_elements(By.TAG_NAME, "caption"):
              checked_tables_caption.append(table)


The_target_table = None
The_target_caption = 'Languages with at least 50 million first-language speakers[7]'

for table in checked_tables_caption:
    elem = table.find_element(By.TAG_NAME, "caption")
    
    if str(elem.text.strip()) == The_target_caption:
        The_target_table = table 
        break

rows = The_target_table.find_elements(By.TAG_NAME, "tr")
The_data = []
The_headers = [h.text.replace('\n', '') for h in The_target_table.find_elements(By.TAG_NAME, 'th')]

for row in rows:
      data = [d.text.replace('\n', '') for d in (row.find_elements(By.TAG_NAME, 'td'))]
      if len(data) == 0:
            continue
      The_data.append(data)

file = openpyxl.Workbook()
sheet = file.active

for header_colomn,header in enumerate(The_headers):
    sheet.cell(row=1, column=header_colomn + 1).value = f'{header}'

for num_line,row in enumerate(The_data):
      for num_unit,data_cell in enumerate(row):
            sheet.cell(row=num_line + 2, column=num_unit + 1).value = f'{data_cell}'
file.save(Path.home() / Path("Desktop", "tests", "SeleniumToXlsx.xlsx"))




