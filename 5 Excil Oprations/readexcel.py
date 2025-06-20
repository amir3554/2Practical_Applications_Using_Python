import openpyxl
from pathlib import Path


# Creat excel file
file_excel = openpyxl.Workbook()
print(file_excel.sheetnames)

# Change excel sheet name and adding sheets
sheet_excel = file_excel.active
sheet_excel.title = 'shatshoota'

file_excel.create_sheet()
file_excel.create_sheet(title='third sheet')
file_excel.create_sheet(title='second sheet', index=1)
file_excel.create_sheet(title='middle sheet', index=2)

# Write on the sheet
sheet2 = file_excel["second sheet"]
sheet2["A1"] = "amoor"
print(sheet2['A1'].value)

# exsersize

file_excel2 = openpyxl.load_workbook(Path.home() / Path("Desktop", 'tests', 'employees.xlsx'))
the_sheet = file_excel2['Sheet1']
the_sheet2 = file_excel["middle sheet"]

for i in range(1, the_sheet.max_row +1):
    employee = (the_sheet.cell(row=i, column=1)).value
    the_sheet2[f'A{i}'] = f'{employee}'

# Save the file
file_excel.save(filename= Path.home() / Path('Desktop', 'tests',) / 'FromPYExcel.xlsx')






