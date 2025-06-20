import openpyxl
from pathlib import Path

file_path = Path.home() / Path('Desktop', 'tests', 'employees.xlsx')
excil_file = openpyxl.load_workbook(file_path)
print(excil_file.sheetnames)
print(excil_file.active)

sheet1 = excil_file["Sheet1"]

print('------------------------------------------------')

print(sheet1['A1'].value)
print(sheet1['A2'].value)
print(sheet1['A3'].value)
print(sheet1['C1'].value)
print(sheet1['A5'].row)
print(sheet1['B1'].column)
print(sheet1['A1'].coordinate)
print(sheet1.cell(row= 6, column=2).value)

print('------------------------------------------------')

for i in range(1, sheet1.max_row + 1):
    print(f'{i} : {sheet1.cell(row=i, column=1).value}')

print('------------------------------------------------')

total = 0

for i in range(1, 7):
    print(f'{i} : {sheet1.cell(row=i, column=1).value} : {sheet1.cell(row=i, column=2).value}')
    total += sheet1.cell(row=i, column=2).value

print(f'the total is :{total}')

